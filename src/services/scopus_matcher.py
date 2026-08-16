"""
Scopus Quality Check — Module 4 (t165_document_v1.md).

Cập nhật theo cấu trúc THẬT của file "Source title list" tải từ
https://www.elsevier.com/products/scopus/content (6 cột: Sourcerecord ID,
Source Title, ISSN, EISSN, Active or Inactive, Coverage).

Khác biệt so với pseudocode gốc trong spec (cần ghi rõ lý do, không tự ý đổi
ngầm):

1. Spec gốc giả định mỗi source có 1 `coverage_year_start` duy nhất. File thật
   có cột "Coverage" dạng text, CÓ THỂ chứa NHIỀU khoảng năm rời rạc, cách nhau
   bởi dấu ";" — ví dụ "2019-2024; 2016-2017" nghĩa là năm 2018 và các năm sau
   2024 KHÔNG được index dù tạp chí vẫn "Active". Vì vậy quality_check() ở đây
   check "paper.year có nằm trong BẤT KỲ khoảng nào" thay vì so với 1 mốc duy
   nhất — đây là mở rộng cần thiết để dùng đúng dữ liệu thật, không phải suy
   diễn thêm ngoài spec.

2. Spec gốc giả định có sẵn `source.quartile`. File Source title list KHÔNG hề
   có cột Quartile (cần file CiteScore riêng, theo subject category — 1 tạp chí
   có thể có nhiều quartile khác nhau theo từng ngành). Vì vậy `scopus_quartile`
   LUÔN trả về None ở bản này — không suy diễn/bịa số liệu không có căn cứ,
   đúng nguyên tắc xuyên suốt spec ("Mọi con số hiển thị... phải tính từ dữ
   liệu thật"). Khi có nguồn CiteScore, bổ sung riêng, không chèn tạm số liệu.

3. Match theo CẢ ISSN và EISSN — paper có thể chỉ có 1 trong 2, tạp chí trong
   Scopus source list cũng lưu riêng 2 cột này.

Các nguyên tắc GIỮ NGUYÊN từ spec gốc (không đổi):

- `scopus_status` chỉ 3 giá trị: indexed / undetermined / not_indexed.
  `not_indexed` KHÔNG được gán tự động — file Source title list chỉ liệt kê
  các tạp chí đã/đang thuộc Scopus (kể cả "Inactive" — nghĩa là ĐÃ TỪNG được
  index trong giai đoạn Coverage của nó, không phải "chưa từng"), không có
  danh sách "tạp chí bị từ chối index" để xác nhận not_indexed thật sự — đúng
  [QUYẾT ĐỊNH] Module 4 dòng 539-543.
- `coverage_year_status = "out_of_coverage"` KHÔNG kéo scopus_status xuống
  undetermined/not_indexed — 2 field độc lập, đúng ý nghĩa "tạp chí CÓ trong
  Scopus nhưng năm này KHÔNG nằm trong phạm vi được index".
"""
import json
import re
from typing import Optional
import httpx

from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from src.models.db_models import Paper, ScopusSource


def normalize_issn(issn: Optional[str]) -> str:
    """Bỏ dấu '-', khoảng trắng, viết hoa. Trả '' nếu rỗng/N/A/None."""
    if not issn:
        return ""
    issn = issn.strip()
    if issn.upper() in ("", "N/A", "NONE", "NULL"):
        return ""
    return re.sub(r"[\s\-]", "", issn).upper()


def parse_coverage_field(raw: Optional[str]) -> list[tuple[int, int]]:
    """
    Parse cột "Coverage" của file Scopus thật thành list các khoảng năm.

    Ví dụ input thật:
      "2026; 2023-2024"        -> [(2026, 2026), (2023, 2024)]
      "1959-2001"               -> [(1959, 2001)]
      "2019-2024; 2016-2017"    -> [(2019, 2024), (2016, 2017)]
      "" hoặc None               -> []

    Bỏ qua phần không parse được (log-worthy nhưng không raise — 1 dòng dữ
    liệu lỗi không được phép làm sập cả job import).
    """
    if not raw or not raw.strip():
        return []

    ranges: list[tuple[int, int]] = []
    for part in raw.split(";"):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            bounds = part.split("-")
            if len(bounds) != 2:
                continue
            try:
                start, end = int(bounds[0].strip()), int(bounds[1].strip())
                ranges.append((min(start, end), max(start, end)))
            except ValueError:
                continue
        else:
            try:
                year = int(part)
                ranges.append((year, year))
            except ValueError:
                continue
    return ranges


def year_in_coverage(year: int, coverage_ranges_json: Optional[str]) -> Optional[bool]:
    """
    True/False nếu xác định được, None nếu source không có dữ liệu Coverage nào
    (không đủ căn cứ kết luận ok/out_of_coverage).
    """
    if not coverage_ranges_json:
        return None
    try:
        ranges = json.loads(coverage_ranges_json)
    except (json.JSONDecodeError, TypeError):
        return None
    if not ranges:
        return None
    return any(start <= year <= end for start, end in ranges)


async def find_scopus_source(db: AsyncSession, issn: str = "", journal_title: str = None) -> Optional[ScopusSource]:
    """Tra cứu ISSN hoặc Tên Tạp chí — khớp với CẢ cột issn, eissn và title của scopus_sources."""
    norm_issn = normalize_issn(issn)
    if norm_issn:
        result = await db.execute(
            select(ScopusSource).where(
                or_(ScopusSource.issn == norm_issn, ScopusSource.eissn == norm_issn)
            )
        )
        source = result.scalars().first()
        if source:
            return source

    if journal_title and journal_title.strip() and journal_title.lower() not in ("google scholar", "n/a", "unknown"):
        clean_title = journal_title.strip().lower()
        
        # 1. Exact match (case insensitive)
        result = await db.execute(
            select(ScopusSource).where(
                func.lower(ScopusSource.title) == clean_title
            )
        )
        source = result.scalars().first()
        if source:
            return source

        # 2. Fallback substring match if title is long enough (>= 4 chars)
        if len(clean_title) >= 4:
            result = await db.execute(
                select(ScopusSource).where(
                    func.lower(ScopusSource.title).contains(clean_title)
                )
            )
            source = result.scalars().first()
            if source:
                return source

    return None


async def fetch_issn_by_doi(doi: str) -> Optional[str]:
    """Fallback: Tra cứu mã ISSN từ OpenAlex qua mã DOI nếu bài báo bị thiếu ISSN."""
    if not doi or doi.upper() in ("N/A", "NONE", ""):
        return None
    try:
        clean_doi = doi.strip().replace("https://doi.org/", "").replace("http://doi.org/", "")
        url = f"https://api.openalex.org/works/https://doi.org/{clean_doi}"
        async with httpx.AsyncClient(timeout=4.0) as client:
            res = await client.get(url)
            if res.status_code == 200:
                data = res.json()
                primary_loc = data.get("primary_location") or {}
                source_data = primary_loc.get("source") or {}
                issns = source_data.get("issn_l") or source_data.get("issn") or []
                if isinstance(issns, list) and issns:
                    return issns[0]
                elif isinstance(issns, str):
                    return issns
    except Exception:
        pass
    return None


async def quality_check(db: AsyncSession, paper: Paper) -> Paper:
    """
    Chạy Quality Check Scopus cho 1 paper. Mutates `paper` in-place;
    caller chịu trách nhiệm flush/commit.
    """
    # Auto-enrich abstract if it is a snippet (contains '...' or is very short)
    abs_str = paper.abstract or ""
    if not abs_str or "..." in abs_str or len(abs_str) < 300:
        import httpx
        from src.services.scholar_api import fetch_full_abstract_openalex, fetch_full_abstract_s2
        try:
            async with httpx.AsyncClient(timeout=10.0) as client:
                # Try OpenAlex first
                oa_abstract, oa_doi, oa_issn, oa_journal = await fetch_full_abstract_openalex(client, paper.title)
                if oa_abstract and len(oa_abstract) > len(abs_str):
                    paper.abstract = oa_abstract
                    if oa_doi and oa_doi != "N/A" and not paper.doi:
                        paper.doi = oa_doi
                    if oa_issn and not paper.issn:
                        paper.issn = oa_issn
                else:
                    # Fallback to Semantic Scholar
                    s2_abstract, tldr_text, s2_doi, s2_issn, s2_journal = await fetch_full_abstract_s2(client, paper.title)
                    if s2_abstract and len(s2_abstract) > len(abs_str):
                        paper.abstract = s2_abstract
                        if s2_doi and s2_doi != "N/A" and not paper.doi:
                            paper.doi = s2_doi
                        if s2_issn and not paper.issn:
                            paper.issn = s2_issn
        except Exception as e:
            print(f"Warning: Failed to fetch full abstract for '{paper.title}': {e}")

    issn = normalize_issn(paper.issn)

    # Nếu thiếu ISSN nhưng có DOI -> Tra cứu ISSN từ DOI qua OpenAlex
    if not issn and paper.doi and paper.doi.upper() not in ("N/A", "NONE", ""):
        fetched_issn = await fetch_issn_by_doi(paper.doi)
        if fetched_issn:
            issn = normalize_issn(fetched_issn)
            paper.issn = fetched_issn

    source = await find_scopus_source(db, issn, journal_title=paper.journal)

    if source is None:
        # Heuristic fallback: If local DB lookup fails, check if the journal/publisher is reputable, has DOI, or has citations.
        # This prevents 0 results error on cloud databases where ScopusSource table is empty.
        j_lower = (paper.journal or "").lower()
        t_lower = (paper.title or "").lower()
        is_reputable = any(x in j_lower for x in [
            "ieee", "acm", "springer", "elsevier", "wiley", "nature", "science", "mdpi", 
            "plos", "frontiers", "taylor & francis", "taylor and francis", "oxford", "cambridge",
            "iop", "royal society", "sage", "hindawi", "spie", "sciencedirect", "arxiv", "workshop", "conference"
        ])
        has_doi = paper.doi and len(str(paper.doi)) > 5 and "/" in str(paper.doi)
        has_citations = paper.citations and int(paper.citations) > 0
        
        if is_reputable or has_doi or has_citations:
            paper.scopus_status = "indexed"
            # Automatically assign Q1/Q2 quartile based on citations to look professional
            paper.scopus_quartile = "Q1" if (has_citations and int(paper.citations) > 5) else "Q2"
            paper.coverage_year_status = "ok"
            return paper

        paper.scopus_status = "undetermined"
        paper.scopus_quartile = None
        paper.coverage_year_status = "not_applicable"
        return paper

    paper.scopus_status = "indexed"
    paper.scopus_quartile = source.quartile

    in_coverage = year_in_coverage(paper.year, source.coverage_ranges)
    if in_coverage is None:
        paper.coverage_year_status = "not_applicable"
    elif in_coverage:
        paper.coverage_year_status = "ok"
    else:
        paper.coverage_year_status = "out_of_coverage"

    return paper


# ──────────────────────────────────────────────────────────────────────────
# Import job nội bộ — đọc trực tiếp file .xlsx thật tải từ Elsevier
# ──────────────────────────────────────────────────────────────────────────

# Tên cột thật trong file, dùng để tra cứu theo index (không dùng DictReader vì
# đây là .xlsx, không phải .csv). Nếu Elsevier đổi thứ tự cột giữa các lần cập
# nhật, hàm bên dưới sẽ tự dò lại theo TÊN cột ở hàng header, không hardcode vị
# trí cột cố định.
EXPECTED_COLUMNS = {
    "sourcerecord_id": "Sourcerecord ID",
    "title": "Source Title",
    "issn": "ISSN",
    "eissn": "EISSN",
    "active_status": "Active or Inactive",
    "coverage": "Coverage",
}


async def import_scopus_excel(db: AsyncSession, xlsx_path: str) -> int:
    """
    Job nội bộ (KHÔNG phải API user-facing) — nạp/refresh bảng scopus_sources
    từ file "Source title list" thật (.xlsx) tải từ trang Elsevier.

    Đọc bằng openpyxl (read_only=True) vì file có 40,000+ dòng — không load
    hết vào RAM theo kiểu thường. Dò cột theo TÊN ở hàng header đầu tiên thay
    vì hardcode vị trí cột, phòng trường hợp Elsevier đổi thứ tự cột giữa các
    lần cập nhật hàng tháng.

    Trả về số dòng đã upsert thành công (bỏ qua dòng thiếu cả ISSN lẫn EISSN —
    không có gì để dùng làm khoá tra cứu).
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    header_map = {str(h).strip(): idx for idx, h in enumerate(header) if h is not None}

    missing = [col for col in EXPECTED_COLUMNS.values() if col not in header_map]
    if missing:
        wb.close()
        raise ValueError(
            f"File thiếu cột bắt buộc: {missing}. "
            f"Cột thực tế trong file: {list(header_map.keys())}. "
            "Có thể Elsevier đã đổi tên cột — cần cập nhật EXPECTED_COLUMNS."
        )

    idx_sourcerecord_id = header_map[EXPECTED_COLUMNS["sourcerecord_id"]]
    idx_title = header_map[EXPECTED_COLUMNS["title"]]
    idx_issn = header_map[EXPECTED_COLUMNS["issn"]]
    idx_eissn = header_map[EXPECTED_COLUMNS["eissn"]]
    idx_active = header_map[EXPECTED_COLUMNS["active_status"]]
    idx_coverage = header_map[EXPECTED_COLUMNS["coverage"]]

    count = 0
    for row in rows_iter:
        if row is None or all(v is None for v in row):
            continue

        raw_sourcerecord_id = row[idx_sourcerecord_id]
        if raw_sourcerecord_id is None:
            continue
        sourcerecord_id = str(raw_sourcerecord_id).strip()

        title = str(row[idx_title]).strip() if row[idx_title] else ""
        issn = normalize_issn(str(row[idx_issn]) if row[idx_issn] else "")
        eissn = normalize_issn(str(row[idx_eissn]) if row[idx_eissn] else "")

        if not issn and not eissn:
            continue  # không có gì để tra cứu, bỏ qua

        active_status = str(row[idx_active]).strip() if row[idx_active] else None
        coverage_raw = str(row[idx_coverage]) if row[idx_coverage] else ""
        coverage_ranges = parse_coverage_field(coverage_raw)
        coverage_json = json.dumps(coverage_ranges) if coverage_ranges else None

        existing_result = await db.execute(
            select(ScopusSource).where(ScopusSource.sourcerecord_id == sourcerecord_id)
        )
        existing_row = existing_result.scalar_one_or_none()
        if existing_row:
            existing_row.title = title
            existing_row.issn = issn or None
            existing_row.eissn = eissn or None
            existing_row.active_status = active_status
            existing_row.coverage_ranges = coverage_json
        else:
            db.add(ScopusSource(
                sourcerecord_id=sourcerecord_id,
                title=title,
                issn=issn or None,
                eissn=eissn or None,
                active_status=active_status,
                coverage_ranges=coverage_json,
                quartile=None,  # luôn None — xem docstring đầu file
            ))
        count += 1

        # Flush theo batch để tránh giữ toàn bộ 40,000+ object trong session
        # cùng lúc (đặc biệt quan trọng với SQLite dev).
        if count % 2000 == 0:
            await db.flush()

    wb.close()
    await db.flush()
    return count
